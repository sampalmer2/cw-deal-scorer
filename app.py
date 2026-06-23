import streamlit as st
import pandas as pd
from scorer import score_property, MODULES, FORMULA_VERSION
from database import (
    init_db, save_property, save_outcome,
    get_all_scores, get_accuracy_metrics, get_disagreements,
    get_outcomes_for_analysis, find_similar_deals, load_all_for_chat,
)
import urllib.parse
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

try:
    import anthropic as _anthropic
    _ANTHROPIC_OK = True
except ImportError:
    _ANTHROPIC_OK = False


def _build_template() -> io.BytesIO:
    """Generate a pre-formatted Excel scoring template with Column Guide sheet."""
    INDIGO     = "1D1740"
    LIGHT_GREY = "F5F6F7"
    WHITE      = "FFFFFF"
    GREEN      = "1D6B28"
    GREY_TEXT  = "6B7280"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scoring Template"

    columns = [
        "Address", "City", "State", "Asset Class", "Zip Code",
        "Annual Rent", "EBITDAR", "Sales",
        "Building SF", "Store Age", "Year Built", "Lot Size (Acres)", "Num Bays",
        "Lease Type", "Lease Term (Years)", "Lease Remaining (Years)", "Lease Expiration",
        "Rent Bumps", "Corporate Guarantee (Y/N)", "Guarantor",
        "AADT", "Pop 1 Mile", "Pop 3 Mile", "Pop 5 Mile",
        "Avg HH Income 1 Mile", "Avg HH Income 3 Mile", "Avg HH Income 5 Mile",
        "MSA", "County", "Market Type",
        "Site Override (-1/0/1/2)", "Access Score (1-5)", "Loc Override (-1/0/1)",
        "Infill Score (1-5)", "Geo Constraint (Y/N)",
        "Brand Tier (1-5)", "Drive Thru Score (1-5)", "Guarantee Score (1-5)",
        "AUV vs Brand (1-5)", "Operator Score (1-5)",
        "Membership Pct (1-5)", "Wash Format (1-5)", "Daily Volume (1-5)",
        "Medical Specialty (1-5)", "Payer Mix (1-5)", "TI Investment (1-5)",
        "Fuel Volume (1-5)", "Inside Sales Pct (1-5)", "Fuel Brand (1-5)",
        "Sales PSF (1-5)", "Lease Structure (1-5)", "Competition Score (1-5)",
        "Membership Penetration (1-5)", "Fitness Format (1-5)", "Equip Lease Alignment (1-5)",
        "Notes", "Caveats",
    ]

    # Generic QSR example — EBITDAR and Sales left blank for blind-mode demo
    example = [
        "123 Main Street", "Austin", "TX", "qsr", "78704",
        325000, None, None,
        3200, 8, 2016, 0.85, None,
        "NNN", 20, 12, "2036-12-31",
        "10% every 5 years", "Y", "Example Tenant LLC",
        38500, 18500, 87000, 195000,
        92000, 88000, 86000,
        "Austin-Round Rock-Georgetown TX", "Travis", "Urban",
        1, 4, 0,
        4, "N",
        5, 4, 5,
        None, None,
        None, None, None,
        None, None, None,
        None, None, None,
        None, None, None,
        None, None, None,
        "Hard corner · signalized intersection · corporate lease · high-traffic corridor",
        "Confirm rent bumps · verify corporate guarantee entity",
    ]

    hdr_fill  = PatternFill("solid", fgColor=INDIGO)
    hdr_font  = Font(color=WHITE, bold=True, size=9, name="Calibri")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ex_fill   = PatternFill("solid", fgColor=LIGHT_GREY)
    ex_font   = Font(size=9, name="Calibri")
    ex_align  = Alignment(horizontal="left", vertical="center", wrap_text=False)

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 18

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = hdr_align

    for col_idx, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill  = ex_fill
        cell.font  = ex_font
        cell.alignment = ex_align

    wide_cols  = {"Address", "MSA", "Notes", "Caveats", "Guarantor", "Rent Bumps", "Asset Class"}
    med_cols   = {"City", "County", "Lease Type", "Market Type", "Lease Expiration",
                  "Avg HH Income 1 Mile", "Avg HH Income 3 Mile", "Avg HH Income 5 Mile"}
    narrow_cols = {"State", "Zip Code", "Num Bays", "Corporate Guarantee (Y/N)",
                   "Geo Constraint (Y/N)"}

    for col_idx, col_name in enumerate(columns, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        if col_name in wide_cols:
            ws.column_dimensions[letter].width = 30
        elif col_name in med_cols:
            ws.column_dimensions[letter].width = 18
        elif col_name in narrow_cols:
            ws.column_dimensions[letter].width = 10
        elif "(1-5)" in col_name or "(-1/0/1" in col_name:
            ws.column_dimensions[letter].width = 16
        else:
            ws.column_dimensions[letter].width = 14

    ws.freeze_panes = "A2"

    # ── Column Guide sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Column Guide")

    for col_idx, label in enumerate(["Column Name", "Description", "Required"], 1):
        cell = ws2.cell(row=1, column=col_idx, value=label)
        cell.fill  = PatternFill("solid", fgColor=INDIGO)
        cell.font  = Font(color=WHITE, bold=True, size=9, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws2.row_dimensions[1].height = 24

    guide = [
        ("Address",                    "Street address of the subject property",                                                                           "No"),
        ("City",                       "City name",                                                                                                         "No"),
        ("State",                      "Two-letter state abbreviation (e.g. TX)",                                                                           "No"),
        ("Asset Class",                "Asset class for scoring. Must be one of: automotive_service, qsr, car_wash, medical, convenience, dollar_store, fitness", "Yes"),
        ("Zip Code",                   "5-digit ZIP code",                                                                                                  "No"),
        ("Annual Rent",                "Annual base rent in dollars",                                                                                       "Yes"),
        ("EBITDAR",                    "Earnings before interest, taxes, D&A and rent. Leave blank to use financial-blind scoring.",                         "No"),
        ("Sales",                      "Unit-level gross revenue. Leave blank to use financial-blind scoring.",                                              "No"),
        ("Building SF",                "Gross building square footage. Default: 12,000",                                                                    "No"),
        ("Store Age",                  "Age of the building in years. Default: 20",                                                                         "No"),
        ("Year Built",                 "Year the building was constructed",                                                                                 "No"),
        ("Lot Size (Acres)",           "Land area in acres",                                                                                                "No"),
        ("Num Bays",                   "Number of service bays (automotive asset class only)",                                                              "No"),
        ("Lease Type",                 "e.g. NNN, Absolute NNN, Modified Gross",                                                                           "No"),
        ("Lease Term (Years)",         "Original lease term length in years",                                                                               "No"),
        ("Lease Remaining (Years)",    "Years remaining on the primary term",                                                                               "No"),
        ("Lease Expiration",           "Lease expiration date in YYYY-MM-DD format",                                                                        "No"),
        ("Rent Bumps",                 "Description of scheduled rent increases (e.g. '10% every 5 years')",                                                "No"),
        ("Corporate Guarantee (Y/N)",  "Y if the lease carries a corporate guarantee, N if personal only",                                                  "No"),
        ("Guarantor",                  "Legal entity providing the lease guarantee",                                                                        "No"),
        ("AADT",                       "Annual Average Daily Traffic count on the fronting road. Leave 0 if unknown.",                                       "No"),
        ("Pop 1 Mile",                 "Total population within a 1-mile radius",                                                                           "No"),
        ("Pop 3 Mile",                 "Total population within a 3-mile radius",                                                                           "No"),
        ("Pop 5 Mile",                 "Total population within a 5-mile radius. Used directly by the scoring formula.",                                     "No"),
        ("Avg HH Income 1 Mile",       "Average household income within a 1-mile radius",                                                                   "No"),
        ("Avg HH Income 3 Mile",       "Average household income within a 3-mile radius",                                                                   "No"),
        ("Avg HH Income 5 Mile",       "Average household income within a 5-mile radius. Used directly by the scoring formula.",                            "No"),
        ("MSA",                        "Metropolitan statistical area name (e.g. 'Austin-Round Rock-Georgetown TX')",                                        "No"),
        ("County",                     "County name",                                                                                                       "No"),
        ("Market Type",                "Urban, Suburban, or Rural",                                                                                         "No"),
        ("Site Override (-1/0/1/2)",   "-1 mid-block/poor visibility · 0 standard pad · +1 hard corner signalized · +2 outparcel anchor-adjacent",          "No"),
        ("Access Score (1-5)",         "Site ingress/egress quality: 1 poor single cut, 3 standard, 5 multiple signalized cuts with turn lane. Default: 3", "No"),
        ("Loc Override (-1/0/1)",      "Trade area context: -1 warehouse/low car ownership · 0 standard · +1 strong co-tenancy or captive corridor",        "No"),
        ("Infill Score (1-5)",         "Supply constraint: 1 greenfield open land · 3 established corridor · 5 irreplaceable infill. Default: 3",           "No"),
        ("Geo Constraint (Y/N)",       "Y if physical geography permanently prevents competitive development (mountain town, coastal peninsula, etc.)",       "No"),
        ("Brand Tier (1-5)",           "QSR / Dollar Store blind-mode: brand strength 1-5. Used when EBITDAR and Sales are not provided.",                   "No"),
        ("Drive Thru Score (1-5)",     "QSR: drive-thru format quality. 5 = double lane or pickup + DT",                                                   "No"),
        ("Guarantee Score (1-5)",      "QSR blind-mode: guarantor credit quality. 5 = corporate investment grade",                                          "No"),
        ("AUV vs Brand (1-5)",         "QSR normal-mode: AUV relative to brand average. 3 = at average, 5 = >130% of brand AUV",                           "No"),
        ("Operator Score (1-5)",       "QSR normal-mode: operator size and track record. 5 = corporate owned",                                              "No"),
        ("Membership Pct (1-5)",       "Car Wash: EFT membership penetration. 5 = >60% membership",                                                        "No"),
        ("Wash Format (1-5)",          "Car Wash: format quality. 1 = self-serve, 5 = express exterior conveyor, new build",                                "No"),
        ("Daily Volume (1-5)",         "Car Wash: cars washed per day. 5 = >300/day average",                                                               "No"),
        ("Medical Specialty (1-5)",    "Medical: specialty and buildout stickiness. 5 = cancer center or hospital affiliate",                               "No"),
        ("Payer Mix (1-5)",            "Medical: payer quality. 1 = >80% Medicaid, 5 = >70% commercial insured",                                           "No"),
        ("TI Investment (1-5)",        "Medical: tenant improvement per SF. 1 = <$50/SF, 5 = >$350/SF hospital-grade",                                      "No"),
        ("Fuel Volume (1-5)",          "Convenience: monthly fuel volume. 1 = <50k gal/month, 5 = >300k gal/month",                                        "No"),
        ("Inside Sales Pct (1-5)",     "Convenience: inside sales quality. 5 = proprietary food program, >30% of revenue",                                  "No"),
        ("Fuel Brand (1-5)",           "Convenience: fuel brand strength. 1 = unbranded, 5 = Wawa/Buc-ee's tier",                                          "No"),
        ("Sales PSF (1-5)",            "Dollar Store: sales per square foot. 1 = <$120/SF, 5 = >$260/SF",                                                   "No"),
        ("Lease Structure (1-5)",      "Dollar Store: expense responsibility. 1 = gross lease, 5 = absolute NNN + corporate guarantee + bumps",             "No"),
        ("Competition Score (1-5)",    "Dollar Store: competitive density. 1 = 2+ competitors within 1 mile, 5 = dominant in underserved market",           "No"),
        ("Membership Penetration (1-5)", "Fitness: active member count. 1 = <500, 5 = >4,000 or wait list",                                                "No"),
        ("Fitness Format (1-5)",       "Fitness: format differentiation. 5 = unique format with high switching cost (boutique, medical fitness)",           "No"),
        ("Equip Lease Alignment (1-5)","Fitness: lease term vs equipment life. 5 = long lease where equipment replacement is prohibitively expensive",      "No"),
        ("Notes",                      "Broker notes — key positives, observations from site visit or deal context",                                        "No"),
        ("Caveats",                    "Items to verify — diligence flags, open questions, conditions to confirm before closing",                            "No"),
    ]

    base_font  = Font(size=9, name="Calibri")
    req_yes    = Font(size=9, name="Calibri", bold=True, color=GREEN)
    req_no     = Font(size=9, name="Calibri", color=GREY_TEXT)
    row_fill   = PatternFill("solid", fgColor=LIGHT_GREY)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, (name, desc, req) in enumerate(guide, 2):
        fill = row_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate([name, desc, req], 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = req_yes if (col_idx == 3 and req == "Yes") else \
                        req_no  if col_idx == 3 else base_font
            cell.alignment = wrap_align
            if fill:
                cell.fill = fill
        ws2.row_dimensions[row_idx].height = 28

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 70
    ws2.column_dimensions["C"].width = 10
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

st.set_page_config(
    page_title="YAFC Deal Scorer",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Global CSS — one injection, no inline HTML elsewhere ───────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', system-ui, -apple-system, sans-serif;
}
.block-container,
[data-testid="stMainBlockContainer"],
[data-testid="stAppViewBlockContainer"] {
    padding-top: 0 !important;
    padding-bottom: 3rem !important;
}
[data-testid="stMetricLabel"] p {
    font-size: 0.62rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.09em !important;
    text-transform: uppercase !important;
    color: #6B7280 !important;
}
[data-testid="stFormSubmitButton"] > button {
    background: #F1B434 !important;
    color: #1D1740 !important;
    font-weight: 600 !important;
    border: none !important;
    border-radius: 4px !important;
    padding: 0.5rem 2rem !important;
}
[data-testid="stFormSubmitButton"] > button:hover {
    background: #d9a028 !important;
}
[data-testid="stDownloadButton"] > button {
    background: transparent !important;
    border: 1.5px solid #1D1740 !important;
    color: #1D1740 !important;
    font-weight: 500 !important;
    border-radius: 4px !important;
    letter-spacing: 0.02em !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: #1D1740 !important;
    color: #fff !important;
}
button[data-baseweb="tab"] {
    font-size: 0.7rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.08em !important;
    text-transform: uppercase !important;
}
h1, h2, h3 {
    color: #1D1740 !important;
    letter-spacing: -0.02em !important;
}
hr { border-color: #EAECEF !important; }
</style>
""", unsafe_allow_html=True)

# ── Header ──────────────────────────────────────────────────────────────────
st.markdown('<div style="padding-top: 1rem"></div>', unsafe_allow_html=True)
st.markdown(
    '<div style="padding:2rem 0 1.5rem 0;">'
    '<p style="font-size:2.2rem;font-weight:700;color:#1D1740;'
    'letter-spacing:-0.03em;margin:0;line-height:1.1;">YAFC Intelligence</p>'
    '<p style="font-size:0.85rem;color:#6B7280;margin:0.4rem 0 0 0;'
    'letter-spacing:0.01em;">'
    'Net Lease Deal Scoring &nbsp;&mdash;&nbsp; Cushman &amp; Wakefield Capital Markets'
    '</p>'
    '</div>',
    unsafe_allow_html=True,
)

# ── Upload parsing helpers ───────────────────────────────────────────────────
def _rint(row, *cols, default=3):
    for c in cols:
        v = row.get(c)
        if v is None:
            continue
        try:
            f = float(v)
            if not (f != f):  # isnan check without importing math
                return int(f)
        except (ValueError, TypeError):
            continue
    return default


def _rfloat(row, *cols, default=0.0):
    for c in cols:
        v = row.get(c)
        if v is None:
            continue
        try:
            f = float(v)
            if not (f != f):
                return f
        except (ValueError, TypeError):
            continue
    return default


def _rbool(row, *cols, default=False):
    for c in cols:
        v = row.get(c)
        if v is None:
            continue
        if isinstance(v, bool):
            return v
        s = str(v).strip().upper()
        if s in ('Y', 'YES', 'TRUE', '1'):
            return True
        if s in ('N', 'NO', 'FALSE', '0'):
            return False
    return default


# Init DB
try:
    init_db()
    db_ok = True
except Exception:
    db_ok = False

tab_score, tab_dash, tab_upload, tab_intel = st.tabs([
    "Score Property", "Dashboard", "Upload Portfolio", "Deal Intelligence",
])

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — SCORE PROPERTY
# ─────────────────────────────────────────────────────────────────────────────
with tab_score:

    # Selectors outside the form so changes re-render inputs
    sel_col, fin_col = st.columns([2, 1])
    with sel_col:
        asset_class = st.selectbox(
            "Asset Class",
            options=list(MODULES.keys()),
            format_func=lambda x: MODULES[x]['name'],
            key="asset_class_selector",
        )
    with fin_col:
        financials_available = st.checkbox(
            "Unit-level financials available",
            value=True,
            key="financials_available",
            help=(
                "Uncheck when you don't have EBITDAR or Sales — "
                "switches S1 to Rent vs Market and S5/S6/S7 to market/brand signals."
            ),
        )
    # Portfolio — outside form so checkbox toggles fields without form submit
    pf_chk_col, pf_name_col, pf_id_col = st.columns([2, 2, 2])
    with pf_chk_col:
        is_portfolio_deal = st.checkbox("Part of a portfolio deal?", key="is_portfolio_deal")
    if is_portfolio_deal:
        with pf_name_col:
            _portfolio_name = st.text_input("Portfolio Name", key="_portfolio_name")
        with pf_id_col:
            _portfolio_id = st.text_input("Portfolio ID (optional)",
                                          placeholder="For grouping runs",
                                          key="_portfolio_id")
    else:
        _portfolio_name = ""
        _portfolio_id   = None

    module_info = MODULES[asset_class]

    # Blind mode variable defaults (overridden by form widgets when shown)
    rent_vs_market          = 0
    brand_strength          = 3
    service_bay_count       = 3
    brand_tier              = 3
    guarantee_score         = 3
    operator_size           = 3
    canopy_condition        = 3
    lease_term_vs_equipment = 3

    with st.form("property_form"):

        scored_by = st.text_input(
            "Scored By",
            placeholder="e.g. Sam Palmer",
            value="",
        )

        # ── Property Details ─────────────────────────────────────────────────
        with st.container(border=True):
            st.markdown("**Property Details**")
            col1, col2 = st.columns(2)
            with col1:
                address     = st.text_input("Address")
                city        = st.text_input("City")
                state       = st.text_input("State")
                annual_rent = st.number_input("Annual Rent ($)",
                                min_value=0, value=350000, step=1000)
                if financials_available:
                    ebitdar = st.number_input("EBITDAR ($)",
                                    min_value=0, value=900000, step=1000)
                else:
                    ebitdar = 0
                    st.caption("Rent vs Market — S1")
                    rent_vs_market = st.select_slider(
                        "Rent vs Market",
                        options=[-2, -1, 0, 1, 2],
                        value=0,
                        format_func=lambda x: {
                            -2: "-2  Well above market — difficult to re-tenant",
                            -1: "-1  Above market rent",
                             0: " 0  At market",
                             1: "+1  Below market — rent upside on renewal",
                             2: "+2  Well below market — significant upside",
                        }[x],
                        label_visibility="collapsed",
                    )
            with col2:
                if financials_available:
                    sales = st.number_input("Sales ($)",
                                    min_value=0, value=4000000, step=10000,
                                    help="Unit-level revenue. Required for Automotive.")
                else:
                    sales = 0
                sf        = st.number_input("Building SF",
                                min_value=0, value=12000, step=100)
                age       = st.number_input("Store Age (years)",
                                min_value=0, value=15, step=1)
                pop_5m    = st.number_input("5-Mile Population",
                                min_value=0, value=75000, step=1000)
                income_5m = st.number_input("5-Mile Median Income ($)",
                                min_value=0, value=95000, step=1000)

        # ── Broker Judgment — Universal (S2, S3, S4) ────────────────────────
        with st.container(border=True):
            st.markdown("**Broker Judgment — Universal Criteria**")
            st.caption(
                "Site visit or Maps check required. "
                "These inputs capture what no database can."
            )

            # S2: Lease Quality — full-width row above the three columns
            lc1, lc2 = st.columns([1, 1])
            with lc1:
                st.caption("Lease Quality — S2")
                lease_score = st.select_slider(
                    "Lease Quality",
                    options=[1, 2, 3, 4, 5],
                    value=3,
                    format_func=lambda x: {
                        1: "1 — <5 yrs remaining or personal guarantee only",
                        2: "2 — 5-7 yrs remaining or weak guarantee",
                        3: "3 — 7-10 yrs or franchisee / subsidiary guarantee",
                        4: "4 — 10-15 yrs, corporate guarantee, standard bumps",
                        5: "5 — 15+ yrs, corporate guarantee, >=10% rent bumps",
                    }[x],
                    label_visibility="collapsed",
                )

            st.divider()

            col3, col4, col5 = st.columns(3)

            with col3:
                st.caption("Physical Asset — S3a")
                site_override = st.select_slider(
                    "Site Geometry",
                    options=[-1, 0, 1, 2],
                    value=0,
                    format_func=lambda x: {
                        -1: "-1  Mid-block, poor visibility",
                         0: " 0  Standard standalone pad",
                         1: "+1  Hard corner, signalized",
                         2: "+2  Outparcel, anchor-adjacent",
                    }[x],
                    label_visibility="collapsed",
                )
                st.caption(
                    "+2 Outparcel: freestanding pad in front of an established anchor. "
                    "Captures road and anchor traffic. Tightest cap rates in the market."
                )

                st.caption("Site Access — S3b")
                access_score = st.select_slider(
                    "Access",
                    options=[1, 2, 3, 4, 5],
                    value=3,
                    format_func=lambda x: {
                        1: "1 — Poor: single cut, no signal, stacking blocks street",
                        2: "2 — Limited: right-in/right-out or shared access",
                        3: "3 — Standard: one or two cuts, acceptable flow",
                        4: "4 — Good: signalized or multiple full-movement cuts",
                        5: "5 — Excellent: multiple cuts, dedicated turn lane",
                    }[x],
                    label_visibility="collapsed",
                )
                st.caption(
                    "Check Street View. Look for curb cuts, signal vs. stop sign, "
                    "stacking room, truck turning radius."
                )

            with col4:
                st.caption("Traffic Count — AADT")
                aadt = st.number_input(
                    "AADT",
                    min_value=0, value=0, step=1000,
                    help="Annual Average Daily Traffic on the fronting road. Leave 0 if unknown.",
                    label_visibility="collapsed",
                )
                if aadt >= 40000:
                    st.success(f"{aadt:,} AADT — major arterial (+1 to S4)")
                elif aadt >= 20000:
                    st.info(f"{aadt:,} AADT — standard corridor (no modifier)")
                elif aadt >= 10000:
                    st.info(f"{aadt:,} AADT — secondary road (no modifier)")
                elif aadt > 0:
                    st.warning(f"{aadt:,} AADT — low traffic (-1 to S4)")

                st.caption("Trade Area Override — S4")
                loc_override = st.select_slider(
                    "Trade Area Override",
                    options=[-1, 0, 1],
                    value=0,
                    format_func=lambda x: {
                        -1: "-1  Warehouse / very low car ownership",
                         0: " 0  Standard",
                         1: "+1  Strong co-tenancy / captive corridor",
                    }[x],
                    label_visibility="collapsed",
                )

            with col5:
                geo_constraint = st.checkbox(
                    "Geographic permanence",
                    value=False,
                    help="Check if physical geography permanently prevents competitive development.",
                )
                if geo_constraint:
                    st.success(
                        "Geographic constraint active — "
                        "S4 weighted for affluent small market. "
                        "S7 Infill floor set to 4 (Automotive)."
                    )
                st.caption(
                    "Mountain towns (BLM / National Forest), resort markets, coastal peninsulas. "
                    "Examples: Sun Valley ID, Jackson WY, Aspen CO."
                )

        # ── Asset Class Module — S5, S6, S7 ─────────────────────────────────
        with st.container(border=True):
            st.markdown(f"**{module_info['name']} Module — S5, S6, S7**")

            if financials_available:
                # ── Normal mode: financial + operational data available ────────
                if asset_class == 'automotive_service':
                    st.caption(
                        "S5 (EBITDAR Margin) and S6 (Store Performance / Rent-to-Sales) "
                        "are calculated from the financial inputs above."
                    )
                    st.caption("Infill & Supply Constraint — S7")
                    infill_score = st.select_slider(
                        "Infill",
                        options=[1, 2, 3, 4, 5],
                        value=3,
                        format_func=lambda x: {
                            1: "1 — Greenfield, open land, no supply constraint",
                            2: "2 — Growth market, land available nearby",
                            3: "3 — Established corridor, standard",
                            4: "4 — Strong infill, limited land, high barriers",
                            5: "5 — Irreplaceable: classic infill or anchor outparcel",
                        }[x],
                        label_visibility="collapsed",
                    )
                    st.caption(
                        "Score 5 if the site cannot be replicated — established urban corridor, "
                        "no available land within 1 mile, or outparcel on an anchor developed decades ago."
                    )
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'qsr':
                    c5, c6, c7 = st.columns(3)
                    with c5:
                        st.caption("AUV vs Brand Average — S5")
                        auv_vs_brand = st.select_slider(
                            "AUV",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — <70% of brand AUV — underperforming unit",
                                2: "2 — 70-90% of brand AUV",
                                3: "3 — 90-110% — at brand average",
                                4: "4 — 110-130% of brand AUV",
                                5: "5 — >130% of brand AUV — top performer",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c6:
                        st.caption("Drive-Thru — S6")
                        drive_thru_score = st.select_slider(
                            "Drive Thru N",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — No drive-thru, format mismatched to market",
                                2: "2 — Walk-up/counter only in drive-thru market",
                                3: "3 — Drive-thru, limited stacking or shared access",
                                4: "4 — Single DT lane, good stacking",
                                5: "5 — Double DT lane or pickup + DT",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c7:
                        st.caption("Operator Quality — S7")
                        operator_score = st.select_slider(
                            "Operator",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Single-unit franchisee, no track record",
                                2: "2 — Small franchisee (<10 units)",
                                3: "3 — Established franchisee (10-50 units)",
                                4: "4 — Top-tier multi-unit franchisee (50+ units)",
                                5: "5 — Corporate owned",
                            }[x],
                            label_visibility="collapsed",
                        )
                    infill_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'car_wash':
                    c5, c6, c7 = st.columns(3)
                    with c5:
                        st.caption("Membership Penetration — S5")
                        membership_pct = st.select_slider(
                            "Membership",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — No membership model, transaction-based only",
                                2: "2 — <20% membership penetration",
                                3: "3 — 20-40% membership",
                                4: "4 — 40-60% membership",
                                5: "5 — >60% EFT membership penetration",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c6:
                        st.caption("Wash Format — S6")
                        wash_format = st.select_slider(
                            "Wash Format N",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Self-serve, minimal investment value",
                                2: "2 — Full service — labor dependent, margin risk",
                                3: "3 — In-bay automatic (IBA)",
                                4: "4 — Express conveyor with detailing add-on",
                                5: "5 — Express exterior conveyor, newer build, high throughput",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c7:
                        st.caption("Daily Volume — S7")
                        daily_volume = st.select_slider(
                            "Volume",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — <50 cars/day — below breakeven",
                                2: "2 — 50-100 cars/day",
                                3: "3 — 100-200 cars/day",
                                4: "4 — 200-300 cars/day",
                                5: "5 — >300 cars/day average",
                            }[x],
                            label_visibility="collapsed",
                        )
                    infill_score = 3
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'medical':
                    c5, c6, c7 = st.columns(3)
                    with c5:
                        st.caption("Specialty & Buildout — S5")
                        medical_specialty = st.select_slider(
                            "Specialty N",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — General/urgent care in generic space",
                                2: "2 — Primary care, basic buildout",
                                3: "3 — Specialty clinic (ortho, cardio), standard TI",
                                4: "4 — Surgery center or high-acuity specialty, heavy TI",
                                5: "5 — Cancer center, hospital affiliate — irreplaceable",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c6:
                        st.caption("Payer Mix — S6")
                        payer_mix = st.select_slider(
                            "Payer Mix",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — >80% Medicaid — high reimbursement risk",
                                2: "2 — Medicaid-heavy, some commercial",
                                3: "3 — Balanced payer mix",
                                4: "4 — Medicare + commercial dominant",
                                5: "5 — >70% commercial, high-income insured market",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c7:
                        st.caption("TI Investment — S7")
                        ti_investment = st.select_slider(
                            "TI N",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — <$50/SF — minimal, generic",
                                2: "2 — $50-100/SF",
                                3: "3 — $100-200/SF — meaningful build-to-suit signal",
                                4: "4 — $200-350/SF — strong anchor investment",
                                5: "5 — >$350/SF — hospital-grade, irreplaceable buildout",
                            }[x],
                            label_visibility="collapsed",
                        )
                    infill_score = 3
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'convenience':
                    c5, c6, c7 = st.columns(3)
                    with c5:
                        st.caption("Fuel Volume — S5")
                        fuel_volume = st.select_slider(
                            "Fuel Volume",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — <50,000 gallons/month",
                                2: "2 — 50-100k gallons/month",
                                3: "3 — 100-200k gallons/month",
                                4: "4 — 200-300k gallons/month",
                                5: "5 — >300k gallons/month",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c6:
                        st.caption("Inside Sales Mix — S6")
                        inside_sales_pct = st.select_slider(
                            "Inside Sales",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Fuel only or <$500k inside sales",
                                2: "2 — Basic convenience, <15% inside margin contribution",
                                3: "3 — Standard c-store, food/beverage mix",
                                4: "4 — Strong foodservice program (Subway, branded)",
                                5: "5 — Proprietary food, loyalty program, >30% inside revenue",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c7:
                        st.caption("Fuel Brand — S7")
                        fuel_brand = st.select_slider(
                            "Fuel Brand N",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Unbranded independent",
                                2: "2 — Regional independent brand",
                                3: "3 — National brand (Shell, BP, Chevron) — standard",
                                4: "4 — National brand with long-term supply agreement",
                                5: "5 — Premium brand (Wawa, Buc-ee's) or exclusive supply",
                            }[x],
                            label_visibility="collapsed",
                        )
                    infill_score = 3
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'dollar_store':
                    c5, c6, c7 = st.columns(3)
                    with c5:
                        st.caption("Sales per SF — S5")
                        sales_psf = st.select_slider(
                            "Sales PSF",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — <$120/SF — below system average",
                                2: "2 — $120-160/SF",
                                3: "3 — $160-210/SF — at system average",
                                4: "4 — $210-260/SF",
                                5: "5 — >$260/SF — high-volume location",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c6:
                        st.caption("Lease Structure — S6")
                        lease_structure = st.select_slider(
                            "Lease Structure N",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Gross lease, landlord bears all costs",
                                2: "2 — Modified gross with partial expenses",
                                3: "3 — NNN with minor landlord carve-outs",
                                4: "4 — Absolute NNN, tenant pays all",
                                5: "5 — Absolute NNN + corporate guarantee + rent bumps",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c7:
                        st.caption("Market Competition — S7")
                        competition_score = st.select_slider(
                            "Competition N",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — 2+ competing dollar stores within 1 mile",
                                2: "2 — 1 direct competitor nearby",
                                3: "3 — Limited local competition",
                                4: "4 — Only dollar store within 3+ miles",
                                5: "5 — Dominant format in underserved market",
                            }[x],
                            label_visibility="collapsed",
                        )
                    infill_score = 3
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'fitness':
                    c5, c6, c7 = st.columns(3)
                    with c5:
                        st.caption("Membership Count — S5")
                        membership_penetration = st.select_slider(
                            "Membership Count",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — <500 active members",
                                2: "2 — 500-1,000 members",
                                3: "3 — 1,000-2,500 members",
                                4: "4 — 2,500-4,000 members",
                                5: "5 — >4,000 active members or wait list",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c6:
                        st.caption("Format & Equipment — S6")
                        fitness_format = st.select_slider(
                            "Fitness Format N",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Basic gym, commodity equipment, no differentiation",
                                2: "2 — Budget gym in highly competitive market",
                                3: "3 — Midmarket — full cardio, weights, group classes",
                                4: "4 — Premium: CrossFit, boutique, or medical fitness",
                                5: "5 — Unique format with high switching cost",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with c7:
                        st.caption("Lease-Equipment Alignment — S7")
                        equip_lease_alignment = st.select_slider(
                            "Lease Alignment",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Short lease, expensive equipment being removed",
                                2: "2 — Lease expires before equipment depreciation",
                                3: "3 — Standard term, acceptable alignment",
                                4: "4 — Lease matches equipment life with renewal options",
                                5: "5 — Long-term lease, equipment replacement cost prohibitive",
                            }[x],
                            label_visibility="collapsed",
                        )
                    infill_score = 3
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = lease_structure = competition_score = 3

            else:
                # ── Financial Blind mode: no unit-level data ──────────────────
                st.caption(
                    "No unit-level financial or operational data — "
                    "using market, brand, and physical signals."
                )

                if asset_class == 'automotive_service':
                    cb5, cb6, cb7 = st.columns(3)
                    with cb5:
                        st.caption("Brand Strength — S5")
                        brand_strength = st.select_slider(
                            "Brand Strength Auto",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Struggling single-market operator",
                                2: "2 — Weak regional presence",
                                3: "3 — Established regional operator",
                                4: "4 — Strong regional brand",
                                5: "5 — Dominant operator (e.g. Les Schwab)",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb6:
                        st.caption("Service Bay Count — S6")
                        service_bay_count = st.select_slider(
                            "Bay Count",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — 1 bay",
                                2: "2 — 2-3 bays",
                                3: "3 — 4-5 bays",
                                4: "4 — 6-7 bays",
                                5: "5 — 8+ bays",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb7:
                        st.caption("Infill & Supply — S7")
                        infill_score = st.select_slider(
                            "Infill Auto Blind",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Greenfield, open land, no supply constraint",
                                2: "2 — Growth market, land available nearby",
                                3: "3 — Established corridor, standard",
                                4: "4 — Strong infill, limited land, high barriers",
                                5: "5 — Irreplaceable: classic infill or anchor outparcel",
                            }[x],
                            label_visibility="collapsed",
                        )
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'qsr':
                    cb5, cb6, cb7 = st.columns(3)
                    with cb5:
                        st.caption("Brand & Market Presence — S5")
                        brand_name = st.selectbox(
                            "Brand",
                            ["Whataburger", "McDonald's", "Chick-fil-A", "Dutch Bros",
                             "Raising Cane's", "Shake Shack", "Five Guys", "Other"],
                            key="qsr_brand_name",
                        )
                        brand_market = st.selectbox(
                            "Brand presence in this market",
                            ["Home market — dominant regional brand",
                             "Strong national presence — top 3 in market",
                             "Standard national presence",
                             "Expansion market — limited locations",
                             "Minimal presence — new or rare in market"],
                            key="qsr_brand_market",
                        )
                        brand_tier = {
                            "Home market — dominant regional brand":    5,
                            "Strong national presence — top 3 in market": 4,
                            "Standard national presence":               3,
                            "Expansion market — limited locations":     2,
                            "Minimal presence — new or rare in market": 1,
                        }[brand_market]
                        st.caption(f"Brand Tier → **{brand_tier} / 5** — brand strength varies by market")
                    with cb6:
                        st.caption("Drive-Thru — S6")
                        drive_thru_score = st.select_slider(
                            "Drive Thru B",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — No drive-thru, format mismatched to market",
                                2: "2 — Walk-up/counter only in drive-thru market",
                                3: "3 — Drive-thru, limited stacking or shared access",
                                4: "4 — Single DT lane, good stacking",
                                5: "5 — Double DT lane or pickup + DT",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb7:
                        st.caption("Guarantee Quality — S7")
                        guarantee_score = st.select_slider(
                            "Guarantee",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Single-unit franchisee, no track record",
                                2: "2 — Small franchisee (<10 units)",
                                3: "3 — Established franchisee (10-50 units)",
                                4: "4 — Top-tier multi-unit (50+ units)",
                                5: "5 — Corporate investment grade",
                            }[x],
                            label_visibility="collapsed",
                        )
                    infill_score = 3
                    auv_vs_brand = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'car_wash':
                    cb5, cb6, cb7 = st.columns(3)
                    with cb5:
                        st.caption("Brand Strength — S5")
                        brand_strength = st.select_slider(
                            "Brand Strength CW",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Independent, no brand recognition",
                                2: "2 — Local brand, limited presence",
                                3: "3 — Regional brand",
                                4: "4 — Strong regional chain",
                                5: "5 — National brand (Mister, Zips, Splash)",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb6:
                        st.caption("Wash Format — S6")
                        wash_format = st.select_slider(
                            "Wash Format B",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Self-serve, minimal investment value",
                                2: "2 — Full service — labor dependent, margin risk",
                                3: "3 — In-bay automatic (IBA)",
                                4: "4 — Express conveyor with detailing add-on",
                                5: "5 — Express exterior conveyor, newer build, high throughput",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb7:
                        st.caption("Infill & Supply — S7")
                        infill_score = st.select_slider(
                            "Infill CW Blind",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Greenfield, open land, no supply constraint",
                                2: "2 — Growth market, land available nearby",
                                3: "3 — Established corridor, standard",
                                4: "4 — Strong infill, limited land, high barriers",
                                5: "5 — Irreplaceable: classic infill or anchor outparcel",
                            }[x],
                            label_visibility="collapsed",
                        )
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'medical':
                    cb5, cb6, cb7 = st.columns(3)
                    with cb5:
                        st.caption("Specialty & Buildout — S5")
                        medical_specialty = st.select_slider(
                            "Specialty B",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Easily relocatable (urgent care in generic space)",
                                2: "2 — Primary care, basic buildout",
                                3: "3 — Specialty clinic, standard TI",
                                4: "4 — Surgery center or heavy-equipment specialty",
                                5: "5 — Equipment-heavy (dental, dialysis) — very sticky",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb6:
                        st.caption("Operator Size — S6")
                        operator_size = st.select_slider(
                            "Operator Size",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Single practice, independent",
                                2: "2 — Small group (<10 locations)",
                                3: "3 — Mid-size group (10-50 locations)",
                                4: "4 — Large regional group (50-100 locations)",
                                5: "5 — National group (100+ locations)",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb7:
                        st.caption("TI Investment — S7")
                        ti_investment = st.select_slider(
                            "TI B",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — <$50/SF — minimal, generic",
                                2: "2 — $50-100/SF",
                                3: "3 — $100-200/SF — meaningful build-to-suit signal",
                                4: "4 — $200-350/SF — strong anchor investment",
                                5: "5 — >$350/SF — hospital-grade, irreplaceable buildout",
                            }[x],
                            label_visibility="collapsed",
                        )
                    infill_score = 3
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    payer_mix = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'convenience':
                    cb5, cb6, cb7 = st.columns(3)
                    with cb5:
                        st.caption("Fuel Brand — S5")
                        fuel_brand = st.select_slider(
                            "Fuel Brand B",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Unbranded independent",
                                2: "2 — Regional independent brand",
                                3: "3 — National brand (Shell, BP, Chevron) — standard",
                                4: "4 — National brand with long-term supply agreement",
                                5: "5 — Premium brand (Wawa, Buc-ee's) or exclusive supply",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb6:
                        st.caption("Canopy Condition — S6")
                        canopy_condition = st.select_slider(
                            "Canopy",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Over 30 years old, deferred maintenance",
                                2: "2 — 20-30 years old, showing age",
                                3: "3 — 10-20 years old, serviceable condition",
                                4: "4 — 5-10 years old, good condition",
                                5: "5 — New build or renovated within 5 years",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb7:
                        st.caption("Infill & Supply — S7")
                        infill_score = st.select_slider(
                            "Infill Conv Blind",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Greenfield, open land, no supply constraint",
                                2: "2 — Growth market, land available nearby",
                                3: "3 — Established corridor, standard",
                                4: "4 — Strong infill, limited land, high barriers",
                                5: "5 — Irreplaceable: classic infill or anchor outparcel",
                            }[x],
                            label_visibility="collapsed",
                        )
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'dollar_store':
                    cb5, cb6, cb7 = st.columns(3)
                    with cb5:
                        st.caption("Brand Tier — S5")
                        brand_tier = st.select_slider(
                            "Brand Tier DS",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Independent dollar concept, no national affiliation",
                                2: "2 — Small regional dollar chain",
                                3: "3 — Established discount retailer",
                                4: "4 — Family Dollar or Dollar Tree",
                                5: "5 — Dollar General corporate",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb6:
                        st.caption("Lease Structure — S6")
                        lease_structure = st.select_slider(
                            "Lease Structure B",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Gross lease, landlord bears all costs",
                                2: "2 — Modified gross with partial expenses",
                                3: "3 — NNN with minor landlord carve-outs",
                                4: "4 — Absolute NNN, tenant pays all",
                                5: "5 — Absolute NNN + corporate guarantee + rent bumps",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb7:
                        st.caption("Market Competition — S7")
                        competition_score = st.select_slider(
                            "Competition B",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — 2+ competing dollar stores within 1 mile",
                                2: "2 — 1 direct competitor nearby",
                                3: "3 — Limited local competition",
                                4: "4 — Only dollar store within 3+ miles",
                                5: "5 — Dominant format in underserved market",
                            }[x],
                            label_visibility="collapsed",
                        )
                    infill_score = 3
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = 3
                    membership_penetration = fitness_format = equip_lease_alignment = 3

                elif asset_class == 'fitness':
                    cb5, cb6, cb7 = st.columns(3)
                    with cb5:
                        st.caption("Brand Strength — S5")
                        brand_strength = st.select_slider(
                            "Brand Strength Fit",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Independent local gym",
                                2: "2 — Small regional fitness chain",
                                3: "3 — Established regional or niche brand",
                                4: "4 — Strong national brand (Anytime, Gold's)",
                                5: "5 — Planet Fitness or Orangetheory national franchise",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb6:
                        st.caption("Format & Equipment — S6")
                        fitness_format = st.select_slider(
                            "Fitness Format B",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Basic gym, commodity equipment, no differentiation",
                                2: "2 — Budget gym in highly competitive market",
                                3: "3 — Midmarket — full cardio, weights, group classes",
                                4: "4 — Premium: CrossFit, boutique, or medical fitness",
                                5: "5 — Unique format with high switching cost",
                            }[x],
                            label_visibility="collapsed",
                        )
                    with cb7:
                        st.caption("Lease Term vs Equipment — S7")
                        lease_term_vs_equipment = st.select_slider(
                            "Lease vs Equipment",
                            options=[1, 2, 3, 4, 5],
                            value=3,
                            format_func=lambda x: {
                                1: "1 — Short lease, expensive equipment recently installed",
                                2: "2 — Lease expires before equipment depreciation",
                                3: "3 — Standard term, acceptable alignment",
                                4: "4 — Lease matches equipment life with renewals",
                                5: "5 — Long lease, equipment replacement prohibitively expensive",
                            }[x],
                            label_visibility="collapsed",
                        )
                    infill_score = 3
                    auv_vs_brand = drive_thru_score = operator_score = 3
                    membership_pct = wash_format = daily_volume = 3
                    medical_specialty = payer_mix = ti_investment = 3
                    fuel_volume = inside_sales_pct = fuel_brand = 3
                    sales_psf = lease_structure = competition_score = 3
                    membership_penetration = equip_lease_alignment = 3

        # ── Market Context ────────────────────────────────────────────────────
        with st.container(border=True):
            st.markdown("**Market Context**")
            st.caption("Macro snapshot at time of scoring — used for portfolio and vintage analysis.")
            mc1, mc2 = st.columns(2)
            with mc1:
                cap_rate_market = st.number_input(
                    "Market Cap Rate (%)",
                    min_value=0.0, max_value=20.0, value=0.0, step=0.05,
                    help="Prevailing cap rate for this asset class in this market. Leave 0 if unknown.",
                )
            with mc2:
                interest_rate_10yr = st.number_input(
                    "10-Year Treasury (%)",
                    min_value=0.0, max_value=20.0, value=0.0, step=0.05,
                    help="Current 10-year Treasury yield at time of scoring. Leave 0 if unknown.",
                )

        # ── Notes, Caveats, Scoring Notes ────────────────────────────────────
        col_n1, col_n2, col_n3 = st.columns(3)
        with col_n1:
            notes = st.text_area(
                "Notes",
                placeholder="Hard corner Spokane WA · outparcel in front of Costco · new build 2022...",
            )
        with col_n2:
            caveats = st.text_area(
                "Caveats",
                placeholder="Confirm lease term · check rent reset · verify anchor tenancy...",
            )
        with col_n3:
            scoring_notes = st.text_area(
                "Scoring Notes",
                placeholder="Formula calibration context, assumptions, or session notes...",
            )

        # ── Broker Thesis ────────────────────────────────────────────────────
        broker_thesis = st.text_area(
            "Broker Thesis",
            placeholder=(
                "Why does this deal trade where it trades? "
                "What does the formula miss? Be specific — "
                "volume, site, market position, tenant story."
            ),
            help=(
                "Why does this deal trade where it trades? "
                "What does the formula miss? Be specific — "
                "volume, site, market position, tenant story."
            ),
            height=150,
        )

        # ── Record type + formula version ────────────────────────────────────
        rt_col, fv_col = st.columns([3, 1])
        with rt_col:
            record_type = st.radio(
                "Record Type",
                ["live", "calibration", "test"],
                index=2,
                horizontal=True,
                help=(
                    "**Live** — real C&W deal, feeds all learning loops. "
                    "**Calibration** — portfolio bid or formula validation, "
                    "excluded from outcome analysis. "
                    "**Test** — dev/scratch, never included in any analysis."
                ),
            )
        with fv_col:
            st.caption("Formula Version")
            st.markdown(f"`{FORMULA_VERSION}`")

        submitted = st.form_submit_button("Score This Property", type="primary")

    # ── Results ───────────────────────────────────────────────────────────────
    if submitted:
        inputs = {
            'asset_class':        asset_class,
            'address':            address,
            'city':               city,
            'state':              state,
            'annual_rent':        annual_rent,
            'ebitdar':            ebitdar,
            'sales':              sales,
            'sf':                 sf,
            'age':                age,
            'pop_5m':             pop_5m,
            'income_5m':         income_5m,
            'lease_score':        lease_score,
            'site_override':      site_override,
            'access_score':       access_score,
            'loc_override':       loc_override,
            'aadt':               aadt,
            'geo_constraint':     geo_constraint,
            # Record metadata
            'record_type':            record_type,
            'formula_version':        FORMULA_VERSION,
            # Financial blind mode flags
            'financials_available':   financials_available,
            'rent_vs_market':         rent_vs_market,
            # Blind mode broker judgment inputs
            'brand_strength':         brand_strength,
            'service_bay_count':      service_bay_count,
            'brand_tier':             brand_tier,
            'guarantee_score':        guarantee_score,
            'operator_size':          operator_size,
            'canopy_condition':       canopy_condition,
            'lease_term_vs_equipment': lease_term_vs_equipment,
            # Module-specific normal mode inputs
            'infill_score':              infill_score,
            'auv_vs_brand':              auv_vs_brand,
            'drive_thru_score':          drive_thru_score,
            'operator_score':            operator_score,
            'membership_pct':            membership_pct,
            'wash_format':               wash_format,
            'daily_volume':              daily_volume,
            'medical_specialty':         medical_specialty,
            'payer_mix':                 payer_mix,
            'ti_investment':             ti_investment,
            'fuel_volume':               fuel_volume,
            'inside_sales_pct':          inside_sales_pct,
            'fuel_brand':                fuel_brand,
            'sales_psf':                 sales_psf,
            'lease_structure':           lease_structure,
            'competition_score':         competition_score,
            'membership_penetration':    membership_penetration,
            'fitness_format':            fitness_format,
            'equip_lease_alignment':     equip_lease_alignment,
            # Session metadata
            'scored_by':                 scored_by or None,
            'is_portfolio_deal':         is_portfolio_deal,
            'portfolio_name':            _portfolio_name or None,
            'portfolio_id':              _portfolio_id or None,
            'cap_rate_market':           cap_rate_market or None,
            'interest_rate_10yr':        interest_rate_10yr or None,
            'scoring_notes':             scoring_notes or None,
            'broker_thesis':             broker_thesis or None,
        }
        result = score_property(inputs)
        # Cache so the display + override section survive checkbox re-renders
        st.session_state['_last_score'] = {
            'result':        result,
            'inputs':        inputs,
            'notes':         notes,
            'caveats':       caveats,
            'scoring_notes': scoring_notes,
        }
        st.session_state.pop('_score_saved', None)

    # ── Results + Override ────────────────────────────────────────────────────
    # Reads from session_state so it persists when the override checkbox reruns.
    _cached = st.session_state.get('_last_score')
    if _cached:
        result    = _cached['result']
        inputs    = _cached['inputs']
        _notes   = _cached['notes']
        _caveats = _cached['caveats']
        _addr     = inputs.get('address', '')
        _city     = inputs.get('city', '')
        _state    = inputs.get('state', '')
        result_ac = result.get('Asset Class', 'automotive_service')

        st.divider()
        if _addr:
            st.markdown(f"### {_addr}, {_city} {_state}")

        # Scoring mode badge
        scoring_mode = result.get('Scoring Mode', 'Financial Blind')
        if scoring_mode == 'Coverage Ratio':
            st.success(f"Scoring Mode: **{scoring_mode}** — EBITDAR/Rent coverage calculated")
        elif scoring_mode == 'Rent/Sales Ratio':
            st.info(f"Scoring Mode: **{scoring_mode}** — Rent as % of sales calculated")
        else:
            st.warning(f"Scoring Mode: **{scoring_mode}** — No unit-level financials; using market/brand signals")

        # Grade, pool, total score
        fine_grade = result.get('Fine Grade', result['Grade'])
        if result_ac == 'qsr':
            g1, g2, g3, g4, _ = st.columns([1, 1, 2, 2, 3])
            g1.metric("Fine Grade",    fine_grade,
                      help="5-tier QSR grade: A, A/B, B, B/C, C")
            g2.metric("Grade (A/B/C)", result['Grade'])
            g3.metric("Pool",        result['Pool'])
            g4.metric("Total Score", f"{result['Total Score']} / 40")
        else:
            g1, g2, g3, _ = st.columns([1, 2, 2, 4])
            g1.metric("Grade",       result['Grade'])
            g2.metric("Pool",        result['Pool'])
            g3.metric("Total Score", f"{result['Total Score']} / 40")

        st.divider()

        # Score breakdown
        st.caption("Score Breakdown")
        score_items = [(k, v) for k, v in result.items() if k.startswith('S') and ' — ' in k]
        s_cols = st.columns(len(score_items))
        for i, (label, val) in enumerate(score_items):
            short = label.split(' —')[0].strip()
            desc  = label.split('— ')[1].strip() if '— ' in label else label
            s_cols[i].metric(short, f"{val} / 5", help=desc)

        st.divider()

        # Financial ratios — conditional on scoring mode
        ebitdar_rent   = result.get('EBITDAR/Rent')
        ebitdar_margin = result.get('EBITDAR Margin')
        rent_sales     = result.get('Rent/Sales')

        if ebitdar_margin is not None and rent_sales is not None:
            st.caption("Financial Ratios")
            m1, m2, m3 = st.columns(3)
            m1.metric("EBITDAR / Rent", f"{ebitdar_rent}x")
            m2.metric("EBITDAR Margin", f"{ebitdar_margin}%")
            m3.metric("Rent / Sales",   f"{rent_sales}%")
        elif ebitdar_rent is not None:
            st.caption("Financial Ratios")
            m1, _ = st.columns([1, 3])
            m1.metric("EBITDAR / Rent", f"{ebitdar_rent}x")

        if _notes:
            st.info(_notes)
        if _caveats:
            st.warning(_caveats)

        if _addr:
            query = urllib.parse.quote(f"{_addr}, {_city}, {_state}")
            st.markdown(
                f"[Google Maps](https://www.google.com/maps/search/{query})"
                f"  ·  "
                f"[Street View](https://www.google.com/maps?q={query}&layer=c)"
            )

        # ── Broker Override ───────────────────────────────────────────────────
        st.divider()
        with st.container(border=True):
            st.markdown("**Broker Override**")
            st.caption(
                "Leave unchecked if you agree with the formula grade — "
                "NULL broker\\_grade = agreement in the database."
            )
            override = st.checkbox("Override formula grade", key="_override_chk")
            broker_grade    = None
            override_reason = None
            override_notes  = None
            if override:
                grade_opts = ['A', 'A/B', 'B', 'B/C', 'C'] if result_ac == 'qsr' else ['A', 'B', 'C']
                broker_grade = st.radio(
                    "Your grade",
                    grade_opts,
                    horizontal=True,
                    key="_broker_grade",
                )
                override_reason = st.selectbox(
                    "Reason for override",
                    ["Credit not captured", "Location premium", "Market knowledge",
                     "Lease structure", "Competition vulnerability", "Redevelopment value",
                     "Relationship dynamics", "Other"],
                    key="_override_reason",
                )
                override_notes = st.text_area(
                    "Override Notes",
                    key="_override_notes",
                    placeholder="Describe why the formula missed this...",
                )
                if override_reason == "Other" and not override_notes:
                    st.warning("Notes are required when reason is 'Other'.")

        # ── Save ─────────────────────────────────────────────────────────────
        st.divider()
        if st.session_state.get('_score_saved'):
            st.success("Saved to database.")
        elif db_ok:
            needs_notes = override and override_reason == "Other" and not override_notes
            if st.button("Save to Database", type="primary", key="_save_btn",
                         disabled=needs_notes):
                try:
                    save_inputs = {
                        **inputs,
                        'broker_grade':    broker_grade,
                        'override_reason': override_reason,
                        'override_notes':  override_notes,
                    }
                    save_property(save_inputs, result, _notes, _caveats)
                    st.session_state['_score_saved'] = True
                    st.rerun()
                except Exception as e:
                    st.warning(f"Database save failed: {e}")
        else:
            st.info(
                "Database not configured. "
                "Add DATABASE_URL to Streamlit secrets to enable saving."
            )

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
with tab_dash:

    if not db_ok:
        st.info("Add DATABASE_URL to Streamlit secrets to enable the dashboard.")
    else:
        st.markdown("## Dashboard")
        st.markdown("*Formula accuracy, partner agreement, and deal outcomes*")
        st.divider()

        try:
            metrics = get_accuracy_metrics()

            if not metrics or metrics.get('total_scored', 0) == 0:
                st.info("No scored properties yet. Score a property and save it to see analytics here.")
            else:
                # ── SCORING SUMMARY ───────────────────────────────────────
                st.markdown("### Scoring Summary")
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("Total Scored",    metrics.get('total_scored', 0))
                c2.metric("Broker Scored",   metrics.get('total_broker_scored', 0))

                broker_scored  = metrics.get('total_broker_scored', 0) or 1
                agreement_rate = round((metrics.get('agreements', 0) / broker_scored) * 100, 1)
                c3.metric("Agreement Rate",  f"{agreement_rate}%")
                c4.metric("Disagreements",   metrics.get('disagreements', 0))
                c5.metric("Closed Deals",    metrics.get('total_with_outcomes', 0))

                st.divider()

                # ── CAP RATE BY GRADE ─────────────────────────────────────
                if metrics.get('total_with_outcomes', 0) > 0:
                    st.markdown("### Cap Rate by Grade")
                    st.caption(
                        "Validates whether formula grades predict market pricing. "
                        "Lower cap rate = formula working correctly."
                    )

                    avg_a = metrics.get('avg_cap_a')
                    avg_b = metrics.get('avg_cap_b')
                    avg_c = metrics.get('avg_cap_c')
                    g1, g2, g3 = st.columns(3)
                    g1.metric(
                        "Grade A — Avg Cap Rate",
                        f"{round(avg_a * 100, 2)}%" if avg_a else "No data yet",
                        help="Should be lowest — A sites command tightest pricing",
                    )
                    g2.metric(
                        "Grade B — Avg Cap Rate",
                        f"{round(avg_b * 100, 2)}%" if avg_b else "No data yet",
                    )
                    g3.metric(
                        "Grade C — Avg Cap Rate",
                        f"{round(avg_c * 100, 2)}%" if avg_c else "No data yet",
                        help="Should be highest — C sites trade widest",
                    )

                    dom_a = metrics.get('avg_dom_a')
                    dom_b = metrics.get('avg_dom_b')
                    dom_c = metrics.get('avg_dom_c')
                    d1, d2, d3 = st.columns(3)
                    d1.metric("Grade A — Avg Days on Market", f"{round(dom_a)}" if dom_a else "No data yet")
                    d2.metric("Grade B — Avg Days on Market", f"{round(dom_b)}" if dom_b else "No data yet")
                    d3.metric("Grade C — Avg Days on Market", f"{round(dom_c)}" if dom_c else "No data yet")

                    st.divider()

                # ── DISAGREEMENT ANALYSIS ─────────────────────────────────
                st.markdown("### Where Formula and Broker Disagree")
                st.caption(
                    "These patterns show where the formula needs refinement. "
                    "The most common override reasons are the next things to fix."
                )
                try:
                    dis_rows, dis_cols = get_disagreements()
                    if dis_rows:
                        df_dis = pd.DataFrame(dis_rows, columns=dis_cols)
                        st.dataframe(df_dis, use_container_width=True)
                        top_state  = df_dis.iloc[0].get('state', 'N/A')
                        top_reason = df_dis.iloc[0].get('reasons', 'N/A')
                        st.info(f"Most common disagreement: **{top_state}** — {top_reason}")
                    else:
                        st.success("Formula and broker agree on all scored properties so far.")
                except Exception as e:
                    st.warning(f"Could not load disagreement data: {e}")

                st.divider()

                # ── OUTCOME ENTRY FORM ────────────────────────────────────
                st.markdown("### Log a Deal Outcome")
                st.caption(
                    "Enter close data after a deal closes. "
                    "This is the most important input for formula calibration."
                )
                try:
                    all_rows, all_cols = get_all_scores()
                    df_all = pd.DataFrame(all_rows, columns=all_cols)

                    if not df_all.empty:
                        df_all['label'] = df_all.apply(
                            lambda r: (
                                f"#{r['id']} — "
                                f"{r.get('address', '?')}, "
                                f"{r.get('city', '?')} "
                                f"{r.get('state', '?')} "
                                f"(Grade {r.get('formula_grade', '?')})"
                            ),
                            axis=1,
                        )
                        selected_label = st.selectbox(
                            "Select property to update",
                            options=df_all['label'].tolist(),
                        )
                        selected_id = int(df_all[df_all['label'] == selected_label]['id'].values[0])

                        with st.form("outcome_form"):
                            st.markdown(f"**Logging outcome for:** {selected_label}")
                            oc1, oc2 = st.columns(2)
                            with oc1:
                                outcome_status = st.selectbox(
                                    "Outcome",
                                    ["Sold", "Under Contract", "Withdrawn", "Still Marketing", "Passed"],
                                )
                                close_date  = st.date_input("Close Date")
                                list_price  = st.number_input("List Price ($)",  min_value=0, value=0, step=10000)
                                close_price = st.number_input("Close Price ($)", min_value=0, value=0, step=10000)
                            with oc2:
                                list_cap    = st.number_input("List Cap Rate",  min_value=0.0, value=0.0, step=0.001, format="%.3f")
                                close_cap   = st.number_input("Close Cap Rate", min_value=0.0, value=0.0, step=0.001, format="%.3f")
                                dom         = st.number_input("Days on Market",    min_value=0, value=0, step=1)
                                num_offers  = st.number_input("Number of Offers",  min_value=0, value=0, step=1)
                                buyer_type  = st.selectbox(
                                    "Buyer Type",
                                    ["1031 Exchange", "Institutional / REIT", "Private / Family Office", "Owner User", "Other"],
                                )
                                financing = st.selectbox("Financing", ["Cash", "Financed", "Assumed"])
                            save_outcome_btn = st.form_submit_button("Save Outcome", type="primary")

                        if save_outcome_btn:
                            save_outcome(selected_id, {
                                'outcome':        outcome_status,
                                'close_date':     str(close_date),
                                'list_price':     list_price  if list_price  > 0 else None,
                                'close_price':    close_price if close_price > 0 else None,
                                'list_cap_rate':  list_cap    if list_cap    > 0 else None,
                                'close_cap_rate': close_cap   if close_cap   > 0 else None,
                                'days_on_market': dom         if dom         > 0 else None,
                                'buyer_type':     buyer_type,
                                'num_offers':     num_offers  if num_offers  > 0 else None,
                                'financing':      financing,
                            })
                            st.success(f"Outcome saved for property #{selected_id}")
                            st.rerun()

                except Exception as e:
                    st.warning(f"Could not load properties for outcome entry: {e}")

                st.divider()

                # ── ALL SCORED PROPERTIES ─────────────────────────────────
                st.markdown("### All Scored Properties")
                try:
                    all_rows, all_cols = get_all_scores()
                    df_scores = pd.DataFrame(all_rows, columns=all_cols)

                    if not df_scores.empty:
                        grade_filter = st.multiselect(
                            "Filter by Formula Grade",
                            options=["A", "B", "C"],
                            default=["A", "B", "C"],
                            key="dashboard_grade_filter",
                        )
                        if 'formula_grade' in df_scores.columns:
                            df_filtered = df_scores[df_scores['formula_grade'].isin(grade_filter)]
                        else:
                            df_filtered = df_scores

                        st.dataframe(df_filtered, use_container_width=True, height=400)

                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_scores.astype(str).to_excel(
                                writer, index=False, sheet_name='Scored Properties'
                            )
                        output.seek(0)
                        st.download_button(
                            "Download All Scores",
                            data=output,
                            file_name="YAFC_All_Scores.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dashboard_download",
                        )
                except Exception as e:
                    st.error(f"Could not load scored properties: {e}")

        except Exception as e:
            st.error(f"Dashboard error: {e}")
            st.caption("Check that DATABASE_URL is set correctly in Streamlit secrets.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — UPLOAD PORTFOLIO
# ─────────────────────────────────────────────────────────────────────────────
with tab_upload:
    st.markdown("### Upload Portfolio")
    st.caption(
        "Upload a spreadsheet with one row per property. "
        "Scores all properties and returns a downloadable graded portfolio."
    )

    with st.expander("Required column names"):
        st.markdown("""
| Column | Example | Notes |
|---|---|---|
| `Address` | 1211 Harrison Ave | |
| `City` | Bellingham | |
| `State` | WA | |
| `Annual Rent` | 349160 | |
| `EBITDAR` | 1531804 | |
| `Sales` | 5807257 | Required for Automotive |
| `Asset Class` | automotive_service | Optional — default automotive_service |
| `Building SF` | 13200 | Optional — default 12,000 |
| `Store Age` | 8 | Optional — default 20 |
| `Pop 5Mi` | 123456 | Optional — default 50,000 |
| `Income 5Mi` | 95000 | Optional — default 90,000 |
| `AADT` | 32000 | Optional — 0 if unknown |
| `Lease Score` | 3 | Optional — 1 to 5 |
| `Site Override` | 0 | Optional — -1, 0, +1, +2 |
| `Access Score` | 3 | Optional — 1 to 5 |
| `Loc Override` | 0 | Optional — -1, 0, +1 |
| `Infill Score` | 3 | Optional — 1 to 5 (Automotive only) |
| `Geo Constraint` | FALSE | Optional |
        """)

    dl_col, sel_col, name_col = st.columns([1, 2, 2])
    with dl_col:
        st.download_button(
            "Download Excel Template",
            data=_build_template(),
            file_name="YAFC_Scoring_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with sel_col:
        upload_asset_class = st.selectbox(
            "Default Asset Class (used for rows without an Asset Class column)",
            options=list(MODULES.keys()),
            format_func=lambda x: MODULES[x]['name'],
            key="upload_asset_class_selector",
        )
    with name_col:
        portfolio_name = st.text_input(
            "Portfolio Name",
            placeholder="e.g. QSR Portfolio 2026 or Automotive Q3",
            key="upload_portfolio_name",
        )

    upload_record_type = st.radio(
        "Record Type for this upload",
        ["live", "calibration", "test"],
        index=1,
        horizontal=True,
        help=(
            "**Calibration** is the right choice for most portfolio uploads. "
            "Choose **Live** only for real C&W active deal flow."
        ),
        key="upload_record_type",
    )

    uploaded_file = st.file_uploader(
        "Excel or CSV file",
        type=["xlsx", "xls", "csv"],
        label_visibility="collapsed",
    )

    if uploaded_file:
        try:
            df = (
                pd.read_csv(uploaded_file)
                if uploaded_file.name.endswith('.csv')
                else pd.read_excel(uploaded_file)
            )
        except Exception as e:
            st.error(f"Could not read file: {e}")
            st.stop()

        st.success(f"{len(df)} properties loaded.")

        results  = []
        errors   = []
        db_saved = 0
        db_fails = 0

        for idx, row in df.iterrows():
            try:
                row_ac = str(row.get('Asset Class', '')).strip()
                prop_ac = row_ac if row_ac in MODULES else upload_asset_class
                prop = {
                    'portfolio_name':  portfolio_name or None,
                    'record_type':     upload_record_type,
                    'formula_version': FORMULA_VERSION,
                    'asset_class':     prop_ac,
                    'address':        str(row.get('Address', '') or ''),
                    'city':           str(row.get('City', '') or ''),
                    'state':          str(row.get('State', '') or ''),
                    'annual_rent':    _rfloat(row, 'Annual Rent', default=0),
                    'ebitdar':        _rfloat(row, 'EBITDAR', default=0),
                    'sales':          _rfloat(row, 'Sales', default=0),
                    'sf':             _rfloat(row, 'Building SF', default=12000),
                    'age':            _rfloat(row, 'Store Age', default=20),
                    'pop_5m':         _rfloat(row, 'Pop 5 Mile', 'Pop 5Mi', default=50000),
                    'income_5m':      _rfloat(row, 'Avg HH Income 5 Mile', 'Income 5Mi', default=90000),
                    'aadt':           _rfloat(row, 'AADT', default=0),
                    'lease_score':     _rint(row,   'Lease Score', default=3),
                    'lease_remaining': _rfloat(row, 'Lease Remaining (Years)', default=0),
                    'site_override':  _rint(row, 'Site Override (-1/0/1/2)', 'Site Override', default=0),
                    'access_score':   _rint(row, 'Access Score (1-5)', 'Access Score', default=3),
                    'loc_override':   _rint(row, 'Loc Override (-1/0/1)', 'Loc Override', default=0),
                    'infill_score':   _rint(row, 'Infill Score (1-5)', 'Infill Score', default=3),
                    'geo_constraint': _rbool(row, 'Geo Constraint (Y/N)', 'Geo Constraint', default=False),
                    # Module-specific scores — all asset classes
                    'brand_tier':             _rint(row, 'Brand Tier (1-5)', default=3),
                    'drive_thru_score':       _rint(row, 'Drive Thru Score (1-5)', default=3),
                    'guarantee_score':        _rint(row, 'Guarantee Score (1-5)', default=3),
                    'auv_vs_brand':           _rint(row, 'AUV vs Brand (1-5)', default=3),
                    'operator_score':         _rint(row, 'Operator Score (1-5)', default=3),
                    'membership_pct':         _rint(row, 'Membership Pct (1-5)', default=3),
                    'wash_format':            _rint(row, 'Wash Format (1-5)', default=3),
                    'daily_volume':           _rint(row, 'Daily Volume (1-5)', default=3),
                    'medical_specialty':      _rint(row, 'Medical Specialty (1-5)', default=3),
                    'payer_mix':              _rint(row, 'Payer Mix (1-5)', default=3),
                    'ti_investment':          _rint(row, 'TI Investment (1-5)', default=3),
                    'fuel_volume':            _rint(row, 'Fuel Volume (1-5)', default=3),
                    'inside_sales_pct':       _rint(row, 'Inside Sales Pct (1-5)', default=3),
                    'fuel_brand':             _rint(row, 'Fuel Brand (1-5)', default=3),
                    'sales_psf':              _rint(row, 'Sales PSF (1-5)', default=3),
                    'lease_structure':        _rint(row, 'Lease Structure (1-5)', default=3),
                    'competition_score':      _rint(row, 'Competition Score (1-5)', default=3),
                    'membership_penetration': _rint(row, 'Membership Penetration (1-5)', default=3),
                    'fitness_format':         _rint(row, 'Fitness Format (1-5)', default=3),
                    'equip_lease_alignment':  _rint(row, 'Equip Lease Alignment (1-5)', default=3),
                }
                scored = score_property(prop)

                # Extract S1–S7 generically
                s_scores = {
                    k.split(' — ')[0]: v
                    for k, v in scored.items()
                    if ' — ' in k and k[0] == 'S'
                }

                row_out = {
                    'Address':      row.get('Address', ''),
                    'City':         row.get('City', ''),
                    'State':        row.get('State', ''),
                    'Asset Class':  prop['asset_class'],
                    'Annual Rent':  prop['annual_rent'],
                    'EBITDAR/Rent': scored['EBITDAR/Rent'],
                }
                row_out.update(s_scores)
                row_out['Total Score'] = scored['Total Score']
                row_out['Grade']       = scored['Grade']
                row_out['Fine Grade']  = scored.get('Fine Grade', scored['Grade'])
                row_out['Pool']        = scored['Pool']

                results.append(row_out)

                if db_ok:
                    try:
                        notes_val   = str(row.get('Notes', '') or '')
                        caveats_val = str(row.get('Caveats', '') or '')
                        save_property(prop, scored, notes_val, caveats_val)
                        db_saved += 1
                    except Exception as db_err:
                        db_fails += 1
                        errors.append(f"Row {idx + 1} DB save failed: {db_err}")

            except Exception as e:
                errors.append(f"Row {idx + 1}: {e}")

        results_df = pd.DataFrame(results)

        if db_ok:
            if db_saved > 0 and db_fails == 0:
                st.success(f"{db_saved} of {len(results_df)} properties saved to database.")
            elif db_saved > 0:
                st.warning(
                    f"{db_saved} saved to database — {db_fails} failed. "
                    "See error details below."
                )
            elif db_fails > 0:
                st.error(f"All {db_fails} database saves failed. Scores still shown below.")

        st.divider()

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total",   len(results_df))
        c2.metric("Grade A", len(results_df[results_df['Grade'] == 'A']))
        c3.metric("Grade B", len(results_df[results_df['Grade'] == 'B']))
        c4.metric("Grade C", len(results_df[results_df['Grade'] == 'C']))

        st.divider()

        grade_filter = st.multiselect(
            "Filter by grade",
            options=["A", "B", "C"],
            default=["A", "B", "C"],
            key="upload_grade_filter",
        )
        filtered = results_df[results_df['Grade'].isin(grade_filter)]

        st.dataframe(filtered, use_container_width=True, height=460)

        st.divider()
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            results_df.to_excel(writer, index=False, sheet_name='Scored Portfolio')
        output.seek(0)
        st.download_button(
            "Download Scored Portfolio",
            data=output,
            file_name="YAFC_Scored_Portfolio.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if errors:
            with st.expander(f"{len(errors)} rows had errors"):
                for e in errors:
                    st.write(e)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — DEAL INTELLIGENCE
# ─────────────────────────────────────────────────────────────────────────────
with tab_intel:
    st.markdown("## Deal Intelligence")
    st.caption("Surface patterns from scored deals and get AI analysis on what the formula misses.")

    if not db_ok:
        st.info("Add DATABASE_URL to Streamlit secrets to enable Deal Intelligence.")
    else:
        # ── SECTION 1 — Similar Deals Lookup ─────────────────────────────────
        st.markdown("### Similar Deals Lookup")

        with st.form("intel_form"):
            fi1, fi2, fi3 = st.columns(3)
            with fi1:
                intel_ac = st.selectbox(
                    "Asset Class",
                    options=list(MODULES.keys()),
                    format_func=lambda x: MODULES[x]['name'],
                    key="intel_ac",
                )
                intel_grade = st.selectbox(
                    "Grade Filter",
                    ["All", "A", "B", "C"],
                    key="intel_grade",
                )
            with fi2:
                intel_state = st.text_input(
                    "State (optional)",
                    placeholder="e.g. TX",
                    key="intel_state",
                )
                intel_rent_min = st.number_input(
                    "Min Annual Rent ($)",
                    min_value=0, value=0, step=10000,
                    key="intel_rent_min",
                )
            with fi3:
                intel_rent_max = st.number_input(
                    "Max Annual Rent ($)",
                    min_value=0, value=0, step=10000,
                    help="Leave 0 for no upper limit.",
                    key="intel_rent_max",
                )

            find_clicked = st.form_submit_button(
                "Find Similar Deals", type="primary"
            )

        if find_clicked:
            try:
                rows = find_similar_deals(
                    asset_class=intel_ac,
                    grade=intel_grade,
                    state=intel_state or None,
                    rent_min=intel_rent_min or None,
                    rent_max=intel_rent_max or None,
                )
                st.session_state['_intel_rows']   = rows
                st.session_state['_intel_ac']     = intel_ac
                st.session_state['_intel_grade']  = intel_grade
                st.session_state.pop('_intel_analysis', None)
            except Exception as e:
                st.error(f"Query failed: {e}")

        _intel_rows = st.session_state.get('_intel_rows')
        if _intel_rows is not None:
            if not _intel_rows:
                st.info("No live deals found matching those filters.")
            else:
                df_intel = pd.DataFrame(_intel_rows)

                # Format for display
                df_display = df_intel.copy()
                df_display['annual_rent']    = df_display['annual_rent'].apply(
                    lambda v: f"${v:,.0f}" if pd.notna(v) else "—"
                )
                df_display['close_cap_rate'] = df_display['close_cap_rate'].apply(
                    lambda v: f"{v*100:.2f}%" if pd.notna(v) else "—"
                )
                df_display['scored_at'] = pd.to_datetime(
                    df_display['scored_at']
                ).dt.strftime('%Y-%m-%d')
                df_display.columns = [
                    "Address", "City", "State", "Asset Class",
                    "Formula Grade", "Broker Grade", "Annual Rent",
                    "Close Cap Rate", "Override Reason", "Broker Thesis", "Scored",
                ]

                st.caption(f"{len(_intel_rows)} deal{'s' if len(_intel_rows) != 1 else ''} found")
                st.dataframe(df_display, use_container_width=True, hide_index=True)

                # ── SECTION 2 — AI Deal Analysis ──────────────────────────────
                st.divider()
                st.markdown("### AI Deal Analysis")

                ai_ready = _ANTHROPIC_OK and ("ANTHROPIC_API_KEY" in st.secrets)

                if not ai_ready:
                    if not _ANTHROPIC_OK:
                        st.warning("anthropic package not installed.")
                    else:
                        st.warning(
                            "Add ANTHROPIC_API_KEY to Streamlit secrets to enable AI analysis."
                        )
                else:
                    if st.button("Analyze with AI", type="primary", key="_intel_analyze"):
                        # Build plain-text table for the prompt
                        prompt_rows = []
                        for r in _intel_rows:
                            cap = f"{r['close_cap_rate']*100:.2f}%" if r.get('close_cap_rate') else "unknown"
                            thesis = r.get('broker_thesis') or "none recorded"
                            override = r.get('override_reason') or "none"
                            broker_g = r.get('broker_grade') or "agreed with formula"
                            prompt_rows.append(
                                f"- {r.get('address','?')}, {r.get('city','?')} {r.get('state','?')} | "
                                f"Asset: {r.get('asset_class','?')} | "
                                f"Formula: {r.get('formula_grade','?')} | "
                                f"Broker: {broker_g} | "
                                f"Rent: ${r.get('annual_rent',0):,.0f} | "
                                f"Cap: {cap} | "
                                f"Override: {override} | "
                                f"Thesis: {thesis}"
                            )
                        table_text = "\n".join(prompt_rows)
                        n = len(_intel_rows)
                        ac_label = MODULES.get(
                            st.session_state.get('_intel_ac', ''), {}
                        ).get('name', st.session_state.get('_intel_ac', ''))
                        grade_label = st.session_state.get('_intel_grade', 'All grades')

                        user_prompt = (
                            f"Here are {n} similar {ac_label} deals "
                            f"({grade_label} grade) our brokers have scored:\n\n"
                            f"{table_text}\n\n"
                            "Based on this deal history, what patterns do you see? "
                            "What factors consistently drove broker overrides? "
                            "What outcomes resulted? "
                            "What should the broker consider when scoring a new deal "
                            "in this category?"
                        )

                        with st.spinner("Analyzing with Claude…"):
                            try:
                                client = _anthropic.Anthropic(
                                    api_key=st.secrets["ANTHROPIC_API_KEY"]
                                )
                                msg = client.messages.create(
                                    model="claude-sonnet-4-6",
                                    max_tokens=1000,
                                    system=(
                                        "You are an expert net lease capital markets analyst "
                                        "for the YAFC team at Cushman & Wakefield. "
                                        "You analyze deal patterns and broker judgment "
                                        "to surface insights."
                                    ),
                                    messages=[{"role": "user", "content": user_prompt}],
                                )
                                analysis = msg.content[0].text
                                st.session_state['_intel_analysis'] = analysis
                            except Exception as e:
                                st.error(f"AI analysis failed: {e}")

                    analysis = st.session_state.get('_intel_analysis')
                    if analysis:
                        st.markdown(
                            f'<div style="background:#F5F6F7;border-left:4px solid #1D1740;'
                            f'padding:1.2rem 1.5rem;border-radius:4px;'
                            f'font-size:0.9rem;line-height:1.7;white-space:pre-wrap;">'
                            f'{analysis}</div>',
                            unsafe_allow_html=True,
                        )

        # ── SECTION 3 — Ask the Database ─────────────────────────────────────
        st.divider()
        st.markdown("### Ask the Database")

        # Load full DB into session_state once per session; refreshable
        if '_chat_db' not in st.session_state:
            try:
                st.session_state['_chat_db'] = load_all_for_chat()
            except Exception as _e:
                st.session_state['_chat_db'] = []
                st.error(f"Could not load database: {_e}")

        _chat_db_rows = st.session_state.get('_chat_db', [])
        n_total = len(_chat_db_rows)

        note_col, refresh_col = st.columns([6, 1])
        with note_col:
            st.caption(
                f"Analyzing **{n_total}** total scored deal{'s' if n_total != 1 else ''} "
                f"across all record types."
            )
        with refresh_col:
            if st.button("Refresh", key="_chat_refresh", use_container_width=True):
                st.session_state.pop('_chat_db', None)
                st.session_state.pop('_chat_history', None)
                st.rerun()

        _ai_ready = _ANTHROPIC_OK and ("ANTHROPIC_API_KEY" in st.secrets)

        if not _ai_ready:
            st.warning(
                "Add `ANTHROPIC_API_KEY` to Streamlit secrets to enable AI chat."
                if _ANTHROPIC_OK else
                "The `anthropic` package is not installed."
            )
        else:
            # Build compact DB string for the system prompt (cap at 300 rows)
            _cap = min(300, n_total)
            _db_lines = []
            for _r in _chat_db_rows[:_cap]:
                _cap_rate = (
                    f"{float(_r['close_cap_rate'])*100:.2f}%"
                    if _r.get('close_cap_rate') else "—"
                )
                _rent = (
                    f"${float(_r['annual_rent']):,.0f}"
                    if _r.get('annual_rent') else "—"
                )
                _thesis = (_r.get('broker_thesis') or "—")[:120]
                _db_lines.append(
                    f"{_r.get('property_name','?')}, {_r.get('city','?')} {_r.get('state','?')} | "
                    f"{_r.get('asset_class','?')} | "
                    f"Grade: {_r.get('formula_grade','?')} | "
                    f"Broker: {_r.get('broker_grade') or 'agreed'} | "
                    f"Rent: {_rent} | Cap: {_cap_rate} | "
                    f"Status: {_r.get('deal_status') or '—'} | "
                    f"Override: {_r.get('override_reason') or '—'} | "
                    f"Thesis: {_thesis} | "
                    f"Type: {_r.get('record_type','?')} | "
                    f"Scored: {str(_r.get('scored_at','?'))[:10]} | "
                    f"By: {_r.get('scored_by') or '—'} | "
                    f"Portfolio: {_r.get('portfolio_name') or '—'}"
                )
            _db_context = "\n".join(_db_lines)
            if n_total > _cap:
                _db_context += f"\n[{n_total - _cap} older records not shown]"

            _system_prompt = (
                "You are an expert net lease capital markets analyst for the "
                "YAFC Leased Investment Team at Cushman & Wakefield. "
                "You have access to the team's proprietary deal scoring database. "
                "Answer questions accurately and concisely. "
                "When referencing specific deals, cite the property name, city, and grade. "
                "If the data doesn't contain enough information to answer, "
                "say so clearly rather than guessing.\n\n"
                f"Current database ({n_total} total records):\n{_db_context}"
            )

            # Chat history
            if '_chat_history' not in st.session_state:
                st.session_state['_chat_history'] = []
            _chat_history = st.session_state['_chat_history']

            # Display last 5 exchanges (10 messages) above the input
            if _chat_history:
                for _msg in _chat_history[-10:]:
                    with st.chat_message(_msg['role']):
                        st.markdown(_msg['content'])

                if st.button("Clear chat", key="_clear_chat"):
                    st.session_state['_chat_history'] = []
                    st.rerun()

            # Input form — clear_on_submit resets the text box after each ask
            with st.form("chat_form", clear_on_submit=True):
                _q_col, _btn_col = st.columns([5, 1])
                with _q_col:
                    _chat_q = st.text_input(
                        "Question",
                        placeholder="e.g. Which QSR deals had broker overrides and why?",
                        label_visibility="collapsed",
                    )
                with _btn_col:
                    _chat_submitted = st.form_submit_button(
                        "Ask", type="primary", use_container_width=True
                    )

            if _chat_submitted and _chat_q.strip():
                _question = _chat_q.strip()

                # Include last 4 exchanges (8 messages) for continuity
                _api_messages = [
                    {"role": _m['role'], "content": _m['content']}
                    for _m in _chat_history[-8:]
                ]
                _api_messages.append({"role": "user", "content": _question})

                with st.spinner("Thinking…"):
                    try:
                        _client = _anthropic.Anthropic(
                            api_key=st.secrets["ANTHROPIC_API_KEY"]
                        )
                        _resp = _client.messages.create(
                            model="claude-sonnet-4-6",
                            max_tokens=1000,
                            system=_system_prompt,
                            messages=_api_messages,
                        )
                        _answer = _resp.content[0].text
                        _chat_history.append({"role": "user",      "content": _question})
                        _chat_history.append({"role": "assistant", "content": _answer})
                        st.session_state['_chat_history'] = _chat_history
                        st.rerun()
                    except Exception as _e:
                        st.error(f"AI request failed: {_e}")
